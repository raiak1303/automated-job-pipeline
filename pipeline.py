# Job Finder v2 - Enhanced Pipeline Orchestrator

import sqlite3
from typing import List, Set
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from new_job_finder import AllJobsFetcher
from job_model import Job
from config import (
    TARGET_LOCATIONS_KEYWORDS, TARGET_JOB_TITLES,
    WORK_TYPES, DB_FILE, OUTPUT_FILE,
    SEND_EMAIL_ON_RUN, INCLUDE_EXCEL_ATTACHMENT
)
from email_notifier import EmailNotifier


class JobFilter:
    """Filter jobs based on criteria"""

    @staticmethod
    def matches_location(location: str) -> bool:
        if not location:
            return False
        location_lower = location.lower()
        return any(keyword in location_lower for keyword in TARGET_LOCATIONS_KEYWORDS)

    @staticmethod
    def matches_job_title(title: str) -> bool:
        if not title:
            return False
        title_lower = title.lower()
        return any(target.lower() in title_lower for target in TARGET_JOB_TITLES)

    @staticmethod
    def matches_work_type(work_type: str) -> bool:
        if not work_type:
            return True
        return any(wt.lower() in work_type.lower() for wt in WORK_TYPES)

    @classmethod
    def filter_jobs(cls, jobs: List[Job]) -> List[Job]:
        filtered = []
        for job in jobs:
            if (cls.matches_location(job.location) and
                    cls.matches_job_title(job.title)):
                filtered.append(job)
        return filtered


class JobPipeline:
    """Main pipeline orchestrator"""

    def __init__(self):
        self.db_file = DB_FILE
        self.output_file = OUTPUT_FILE
        self.jobs: List[Job] = []
        self.init_database()

    def init_database(self):
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                company TEXT NOT NULL,
                location TEXT NOT NULL,
                work_type TEXT,
                url TEXT UNIQUE,
                apply_url TEXT,
                company_website TEXT,
                recruiter_name TEXT,
                recruiter_email TEXT,
                recruiter_url TEXT,
                description TEXT,
                source TEXT,
                posted_date TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.commit()
        conn.close()

    def fetch_jobs(self) -> List[Job]:
        print("\n" + "="*60)
        print("FETCHING JOBS FROM ALL SOURCES")
        print("="*60)
        fetcher = AllJobsFetcher()
        self.jobs = fetcher.fetch_all()
        print(f"\nTotal jobs fetched: {len(self.jobs)}")
        return self.jobs

    def filter_jobs(self) -> List[Job]:
        print("\n" + "="*60)
        print("FILTERING JOBS")
        print("="*60)
        before = len(self.jobs)
        self.jobs = JobFilter.filter_jobs(self.jobs)
        print(f"Before : {before} | After : {len(self.jobs)} | Removed : {before - len(self.jobs)}")
        return self.jobs

    def deduplicate_jobs(self) -> List[Job]:
        print("\n" + "="*60)
        print("DEDUPLICATING")
        print("="*60)
        seen: Set[str] = set()
        unique = []
        for job in self.jobs:
            key = f"{job.title.lower()}:{job.company.lower()}:{job.location.lower()}"
            if key not in seen:
                seen.add(key)
                unique.append(job)
        removed = len(self.jobs) - len(unique)
        self.jobs = unique
        print(f"Removed {removed} duplicates | Unique jobs: {len(self.jobs)}")
        return self.jobs

    def save_to_database(self):
        print("\n" + "="*60)
        print("SAVING TO DATABASE")
        print("="*60)
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        saved = 0
        for job in self.jobs:
            try:
                cursor.execute('''
                    INSERT INTO jobs (
                        title, company, location, work_type, url, apply_url,
                        company_website, recruiter_name, recruiter_email,
                        recruiter_url, description, source, posted_date
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    job.title, job.company, job.location, job.work_type,
                    job.url, job.apply_url, job.company_website,
                    job.recruiter_name, job.recruiter_email, job.recruiter_url,
                    job.description[:300] if job.description else '',
                    job.source, job.posted_date
                ))
                saved += 1
            except sqlite3.IntegrityError:
                pass
        conn.commit()
        conn.close()
        print(f"Saved {saved} new jobs to database")

    def export_to_excel(self):
        print("\n" + "="*60)
        print("EXPORTING TO EXCEL")
        print("="*60)
        if not self.jobs:
            print("No jobs to export!")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Job Opportunities"

        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        headers = [
            "Job Title", "Company", "Location", "Work Type",
            "Apply URL", "Company Website",
            "Recruiter Name", "Recruiter Email", "Recruiter LinkedIn",
            "Source", "Posted Date"
        ]

        col_widths = [30, 25, 22, 15, 40, 35, 22, 30, 35, 15, 15]

        # Header row
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 30

        # Data rows
        for row_idx, job in enumerate(self.jobs, 2):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)

            row_data = [
                job.title,
                job.company,
                job.location,
                job.work_type,
                job.apply_url or job.url,
                job.company_website,
                job.recruiter_name,
                job.recruiter_email,
                job.recruiter_url,
                job.source,
                job.posted_date or "N/A",
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value or "")
                cell.border = border
                cell.alignment = left
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

            ws.row_dimensions[row_idx].height = 18

        # Freeze header row
        ws.freeze_panes = "A2"

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Job Search Summary"
        ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws2["A3"] = "Generated:"
        ws2["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws2["A4"] = "Total Jobs:"
        ws2["B4"] = len(self.jobs)
        ws2["A6"] = "Jobs by Source:"
        ws2["A6"].font = Font(bold=True)

        source_counts = {}
        for job in self.jobs:
            source_counts[job.source] = source_counts.get(job.source, 0) + 1

        for i, (source, count) in enumerate(sorted(source_counts.items(), key=lambda x: -x[1]), 7):
            ws2[f"A{i}"] = source
            ws2[f"B{i}"] = count

        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 15

        wb.save(self.output_file)
        print(f"✓ Exported {len(self.jobs)} jobs to {self.output_file}")

    def send_email_notification(self):
        if not SEND_EMAIL_ON_RUN:
            return
        print("\n" + "="*60)
        print("SENDING EMAIL")
        print("="*60)
        notifier = EmailNotifier()
        excel_file = self.output_file if INCLUDE_EXCEL_ATTACHMENT else None
        notifier.send_jobs_report(self.jobs, excel_file)

    def run(self):
        print(f"\n{'='*60}")
        print(f"JOB FINDER v2 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}")
        self.fetch_jobs()
        self.filter_jobs()
        self.deduplicate_jobs()
        self.save_to_database()
        self.export_to_excel()
        self.send_email_notification()
        print(f"\n{'='*60}")
        print(f"COMPLETE — {len(self.jobs)} jobs ready in {self.output_file}")
        print(f"{'='*60}\n")


if __name__ == "__main__":
    pipeline = JobPipeline()
    pipeline.run()
